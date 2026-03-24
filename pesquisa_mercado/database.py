import sqlite3
import pandas as pd
from datetime import datetime
import logging
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

DB_PATH = "historico_pricing.db"

def get_connection():
    """Retorna uma conexão ativa com o banco SQLite local."""
    return sqlite3.connect(DB_PATH)

def _garantir_colunas(conn, table_name: str, df: pd.DataFrame):
    """Garante que a tabela tenha todas as colunas do df antes do append do pandas."""
    cursor = conn.cursor()
    cursor.execute(f"PRAGMA table_info({table_name})")
    columns_info = cursor.fetchall()
    
    if columns_info:
        existing_cols = [info[1] for info in columns_info]
        for col in df.columns:
            if col not in existing_cols:
                logger.info(f"Adicionando coluna '{col}' na tabela {table_name}")
                try:
                    cursor.execute(f'ALTER TABLE {table_name} ADD COLUMN "{col}" TEXT')
                except Exception as e:
                    logger.warning(f"Aviso ao alterar schema de {table_name}: {e}")

def salvar_mercado(df_raw: pd.DataFrame, df_resumo: pd.DataFrame) -> bool:
    """Consolida os dados puros de extração das SERPs no banco local."""
    if df_raw is None or df_raw.empty or df_resumo is None or df_resumo.empty:
        return False
        
    try:
        # Extrações trancadas por Dia (Visão Snapshot Semanal)
        batch_ts = datetime.now().strftime("%Y-%m-%d")
        
        df_raw_save = df_raw.copy()
        df_raw_save["batch_timestamp"] = batch_ts
        
        df_resumo_save = df_resumo.copy()
        df_resumo_save["batch_timestamp"] = batch_ts
        
        # Inserção Atômica
        with get_connection() as conn:
            _garantir_colunas(conn, "Fato_Mercado", df_raw_save)
            df_raw_save.to_sql("Fato_Mercado", conn, if_exists="append", index=False)
            
            _garantir_colunas(conn, "Dim_Regra_GBB", df_resumo_save)
            df_resumo_save.to_sql("Dim_Regra_GBB", conn, if_exists="append", index=False)
            
        logger.info(f"Salvo batch {batch_ts} no BD com {len(df_raw_save)} linhas cruas.")
        return True
    except Exception as e:
        logger.error(f"Erro ao salvar Mercado no BD ODS: {e}")
        return False


def salvar_cruzamento(df_cruzamento: pd.DataFrame) -> bool:
    """Salva a fact table purificada B2B para rastreamento longitudinal de markups."""
    if df_cruzamento is None or df_cruzamento.empty:
        return False
        
    try:
        batch_ts = datetime.now().strftime("%Y-%m-%d")
        df_save = df_cruzamento.copy()
        df_save["batch_timestamp"] = batch_ts
        
        with get_connection() as conn:
            _garantir_colunas(conn, "Fato_Cruzamento", df_save)
            df_save.to_sql("Fato_Cruzamento", conn, if_exists="append", index=False)
            
        logger.info(f"Kpis B2B Cruzados no batch {batch_ts} foram persistidos no BD.")
        return True
    except Exception as e:
        logger.error(f"Erro ao salvar Fato Cruzamento no ODS: {e}")
        return False


def formatar_aba_excel(ws, df: pd.DataFrame):
    """Aplica o padrão C-Level visual na Exportação Histórica."""
    if df.empty: return
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    max_col = get_column_letter(df.shape[1])
    ws.auto_filter.ref = f"A1:{max_col}{len(df)+1}"
    
    colunas_moeda = [
        "preco", "preco_original", "Seu Preço", "Seu Preço Médio BI",
        "Média Global do Mercado", "Média da Sua Faixa (Target)", 
        "Preço Mínimo", "Preço Médio Mercado", "Preço Máximo", 
        "Corte P33 (Better)", "Corte P67 (Best)",
        "Média Econômica", "Média Intermediária", "Média Premium"
    ]
    
    for idx, col in enumerate(df.columns, 1):
        col_series = df[col].astype(str)
        max_len = max(max(col_series.map(len)) if not col_series.empty else 10, len(str(col))) + 2
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len, 40)
        
        if col in colunas_moeda:
            for r in range(2, len(df) + 2):
                ws.cell(row=r, column=idx).number_format = '"R$ "#,##0.00'
        elif str(col) == "batch_timestamp":
            for r in range(2, len(df) + 2):
                ws.cell(row=r, column=idx).number_format = 'dd/mm/yyyy'

def _limpar_df_para_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Força as colunas numéricas para float a fim de não exportar strings com ponto como se fossem texto."""
    colunas_moeda = [
        "preco", "preco_original", "Seu Preço", "Seu Preço Médio BI",
        "Média Global do Mercado", "Média da Sua Faixa (Target)", 
        "Preço Mínimo", "Preço Médio Mercado", "Preço Máximo", 
        "Corte P33 (Better)", "Corte P67 (Best)",
        "Média Econômica", "Média Intermediária", "Média Premium"
    ]
    if "batch_timestamp" in df.columns:
        df["batch_timestamp"] = pd.to_datetime(df["batch_timestamp"], format='mixed', errors='coerce').dt.date
    for col in colunas_moeda:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df

def exportar_banco_completo():
    """Lê todas as tabelas históricas do SQL e empacota em um arquivo binário Excel."""
    import io
    try:
        with get_connection() as conn:
            tabelas = pd.read_sql("SELECT name FROM sqlite_master WHERE type='table';", conn)
            tabelas_nomes = tabelas['name'].tolist()
            
            tabelas_alvo = [t for t in tabelas_nomes if t in ["Fato_Cruzamento", "Fato_Mercado", "Dim_Regra_GBB"]]
            if not tabelas_alvo:
                return None
                
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                if "Fato_Cruzamento" in tabelas_nomes:
                    df = pd.read_sql("SELECT * FROM Fato_Cruzamento", conn)
                    df = _limpar_df_para_excel(df)
                    df.to_excel(writer, index=False, sheet_name="Visão_Analítica (Kpis)")
                    formatar_aba_excel(writer.sheets["Visão_Analítica (Kpis)"], df)
                if "Fato_Mercado" in tabelas_nomes:
                    df = pd.read_sql("SELECT * FROM Fato_Mercado", conn)
                    df = _limpar_df_para_excel(df)
                    df.to_excel(writer, index=False, sheet_name="Raw_Mercado (Concorrentes)")
                    formatar_aba_excel(writer.sheets["Raw_Mercado (Concorrentes)"], df)
                if "Dim_Regra_GBB" in tabelas_nomes:
                    df = pd.read_sql("SELECT * FROM Dim_Regra_GBB", conn)
                    df = _limpar_df_para_excel(df)
                    df.to_excel(writer, index=False, sheet_name="Dim_Banda_Estatistica")
                    formatar_aba_excel(writer.sheets["Dim_Banda_Estatistica"], df)
                    
            return output.getvalue()
    except Exception as e:
        logger.error(f"Erro ao exportar DB completo: {e}")
        return None
