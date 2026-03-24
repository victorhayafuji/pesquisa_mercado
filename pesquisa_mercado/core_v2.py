import json
import logging
import math
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter

from config import N_PAGINAS_DEFAULT, OUTPUT_DIR
from google_shopping_client import GoogleShoppingError, buscar_google_shopping
from processador_resultados import extrair_resultados_basicos

logger = logging.getLogger(__name__)


def carregar_regras(filepath: str = "regras_limpeza.json") -> Dict[str, Any]:
    """Carrega o JSON de regras de limpeza e exclusão."""
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        logger.warning(f"Arquivo {filepath} não encontrado. Regras vazias serão usadas.")
        return {}


def coletar_dados_brutos(palavra_chave: str, n_paginas: int = N_PAGINAS_DEFAULT) -> pd.DataFrame:
    """Extrai resultados brutos via API e padroniza o tipo dos dados."""
    todos_resultados = []
    for page in range(1, n_paginas + 1):
        try:
            resposta_json = buscar_google_shopping(q=palavra_chave, page=page)
            resultados_pagina = extrair_resultados_basicos(resposta_json, palavra_chave)
            for r in resultados_pagina:
                r["pagina"] = page
            todos_resultados.extend(resultados_pagina)
            if not resultados_pagina:
                break
        except Exception as e:
            logger.error(f"Erro na página {page}: {str(e)}")
            break

    if not todos_resultados:
        return pd.DataFrame()

    df = pd.DataFrame(todos_resultados)

    # Deduplicar base por product_id e (produto, seller, preco)
    if "product_id" in df.columns:
        df = df.drop_duplicates(subset=["product_id"], keep="first")
    if set(["produto", "seller", "preco"]).issubset(df.columns):
        df = df.drop_duplicates(subset=["produto", "seller", "preco"], keep="first")

    # Cast numérico
    if "preco" in df.columns:
        df["preco"] = pd.to_numeric(df["preco"], errors="coerce")
    if "rating" in df.columns:
        df["rating"] = pd.to_numeric(df["rating"], errors="coerce")
    if "reviews" in df.columns:
        df["reviews"] = pd.to_numeric(df["reviews"], errors="coerce")

    return df


def build_offtopic_regex(regras: Dict) -> re.Pattern:
    """Monta o regex global de termos a serem excluídos."""
    globais = regras.get("filtros_globais", {}).get("termos_exclusao_regex", [])
    if not globais:
        return re.compile(r"a^") # nunca da match
    return re.compile("|".join(globais), re.IGNORECASE)


def aplicar_limpeza(df: pd.DataFrame, regras: Dict, palavra_chave: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Applica a regra de negócio GBB para filtrar dados inúteis."""
    if df.empty or "produto" not in df.columns:
        return df.copy(), pd.DataFrame()

    remove_mask = pd.Series(False, index=df.index)
    
    # Detecção de Duplicidade (Mesmo Vendedor e Mesmo Preço)
    if "seller" in df.columns and "preco" in df.columns:
        mask_valida = df["seller"].notna() & (df["seller"] != "") & df["preco"].notna()
        # Encontra as duplicatas apenas no subconjunto válido
        duplicatas_idx = df[mask_valida].duplicated(subset=["seller", "preco"], keep="first")
        # Marca True na máscara global pros índices que deram duplicação
        remove_mask.loc[duplicatas_idx[duplicatas_idx].index] = True
    
    offtopic_regex = build_offtopic_regex(regras)
    
    kw_norm = palavra_chave.lower().strip()
    regras_kw = regras.get("regras_por_palavra_chave", {}).get(kw_norm, {})
    termos_obrigatorios = regras_kw.get("termos_obrigatorios_ou_de_seguranca", [])
    termos_exclusao_adc = regras_kw.get("termos_exclusao_adicionais", [])
    
    ctrl = regras.get("controle_incoerencia_preco", {})
    limite_alto = ctrl.get("limite_preco_alto_para_itens_pequenos", 10000)
    termos_pequenos = ctrl.get("termos_item_pequeno", [])
    termos_lote = ctrl.get("termos_item_em_lote", [])

    for i, row in df.iterrows():
        prod = str(row.get("produto", "")).lower()
        preco = row.get("preco", np.nan)

        # Tratar regras da palavra-chave
        if offtopic_regex.search(prod):
            if termos_obrigatorios and any(t in prod for t in termos_obrigatorios):
                pass # é seguro manter
            else:
                remove_mask.at[i] = True

        if termos_exclusao_adc and any(t in prod for t in termos_exclusao_adc):
            remove_mask.at[i] = True
            
        # Incoerência de preço
        if pd.notna(preco) and preco >= limite_alto:
            is_pequeno = any(t in prod for t in termos_pequenos)
            is_lote = any(t in prod for t in termos_lote)
            if is_pequeno and not is_lote:
                remove_mask.at[i] = True
                
    # Filtro IQ (Tukey) para Preços Extremamente Aberrantes
    if df["preco"].notna().sum() > 10:
        p25 = df["preco"].quantile(0.25)
        p75 = df["preco"].quantile(0.75)
        iqr = p75 - p25
        limite_superior_super_extremo = p75 + (5.0 * iqr) # Apenas super discrepantes da base central
        remove_mask = remove_mask | (df["preco"] > limite_superior_super_extremo)

    df_tratado = df.loc[~remove_mask].copy()
    df_removidos = df.loc[remove_mask].copy()

    return df_tratado, df_removidos


def calcular_gbb(df: pd.DataFrame, min_itens_relevantes: int = 10) -> Dict[str, Any]:
    """Calcula estatística descritiva e enquadramento Good/Better/Best."""
    if df.empty or "preco" not in df.columns or df["preco"].notna().sum() < min_itens_relevantes:
        return {}

    precos = df["preco"].dropna()
    p33 = float(precos.quantile(0.33))
    p67 = float(precos.quantile(0.67))

    return {
        "amostra_base": int(len(precos)),
        "minimo": float(precos.min()),
        "media": float(precos.mean()),
        "media_economica": float(precos[precos <= p33].mean()) if not precos[precos <= p33].empty else float(p33),
        "media_intermediaria": float(precos[(precos > p33) & (precos <= p67)].mean()) if not precos[(precos > p33) & (precos <= p67)].empty else float((p33+p67)/2),
        "media_premium": float(precos[precos > p67].mean()) if not precos[precos > p67].empty else float(p67),
        "maximo": float(precos.max()),
        "corte_p33": p33,
        "corte_p67": p67,
    }


def classificar_faixa_gbb(preco: float, p33: float, p67: float) -> str:
    """Aplica o label em um preco individual."""
    if pd.isna(preco):
        return "N/A"
    if preco <= p33:
        return "Econômico (Good)"
    elif preco <= p67:
        return "Médio (Better)"
    return "Premium (Best)"


def exportar_planilha_gbb(df_tratado: pd.DataFrame, gbb_info: Dict, palavra_chave: str, filepath: Path) -> bool:
    """Gera um Excel consolidado com Dados e Regras GBB."""
    try:
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        # Enriquecer o tratado com a banda GBB
        if gbb_info:
            df_tratado["Banda_GBB"] = df_tratado["preco"].apply(
                lambda p: classificar_faixa_gbb(p, gbb_info["corte_p33"], gbb_info["corte_p67"])
            )

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df_tratado.to_excel(writer, index=False, sheet_name="Dados Tratados")
            
            # Formatar a largura e número na aba Dados
            ws_dados = writer.sheets["Dados Tratados"]
            for idx, col in enumerate(df_tratado.columns, 1):
                ws_dados.column_dimensions[get_column_letter(idx)].width = 20
                if col in ["preco", "preco_original"]:
                    for r in range(2, len(df_tratado) + 2):
                        ws_dados.cell(row=r, column=idx).number_format = "#,##0.00"

            # Aba de resumo Régua de Preço
            if gbb_info:
                df_resumo = pd.DataFrame([{
                    "Palavra-Chave": palavra_chave,
                    "Volume Base": gbb_info["amostra_base"],
                    "Preço Mínimo": gbb_info["minimo"],
                    "Preço Médio": gbb_info["media"],
                    "Preço Máximo": gbb_info["maximo"],
                    "Corte Good-Better (P33)": gbb_info["corte_p33"],
                    "Corte Better-Best (P67)": gbb_info["corte_p67"],
                    "Faixa Econômica (Good)": f"Até R$ {gbb_info['corte_p33']:,.2f}",
                    "Faixa Média (Better)": f"R$ {gbb_info['corte_p33']:,.2f} a R$ {gbb_info['corte_p67']:,.2f}",
                    "Faixa Premium (Best)": f"Acima de R$ {gbb_info['corte_p67']:,.2f}",
                }])
                df_resumo.to_excel(writer, index=False, sheet_name="Régua GBB")
                ws_resumo = writer.sheets["Régua GBB"]
                for i in range(1, 11):
                    ws_resumo.column_dimensions[get_column_letter(i)].width = 25
                    
        return True
    except Exception as e:
        logger.error(f"Erro ao exportar planilha GBB: {e}")
        return False
