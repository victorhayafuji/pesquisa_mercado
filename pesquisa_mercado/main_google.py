# main.py

from typing import List, Dict, Any, Tuple
from datetime import datetime
import math
import re

import pandas as pd

from config import OUTPUT_DIR, N_PAGINAS_DEFAULT

# Saída padronizada para análise (padrão Brasil)
ANALISE_DIR = OUTPUT_DIR.parent / "outputs_analise"
from google_shopping_client import buscar_google_shopping, GoogleShoppingError
from processador_resultados import extrair_resultados_basicos
from pathlib import Path
def _slugify(texto: str) -> str:
    """Gera um identificador seguro para nome de arquivo."""
    texto = (texto or "").strip().lower()
    texto = re.sub(r"\s+", "_", texto)
    texto = re.sub(r"[^a-z0-9_\-]+", "", texto)
    return texto[:80] or "busca"


def _exportar_xlsx_ptbr(df: pd.DataFrame, caminho_xlsx: Path) -> bool:
    """
    Exporta o DataFrame para XLSX aplicando formatação numérica amigável ao padrão brasileiro
    (a exibição final depende das configurações regionais do Excel/SO, mas o arquivo sai com
    number_format apropriado para valores numéricos).
    """
    try:
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Aviso: openpyxl não está instalado. Para gerar XLSX, execute: pip install openpyxl")
        return False

    caminho_xlsx.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(caminho_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="dados")
        ws = writer.sheets["dados"]

        headers = [cell.value for cell in ws[1]]
        header_to_col = {h: idx + 1 for idx, h in enumerate(headers) if h}

        # Formatos numéricos (Excel aplica separadores conforme a localidade)
        fmt_int = "#,##0"
        fmt_2 = "#,##0.00"
        fmt_3 = "#,##0.000"

        cols_int = ["posicao", "pagina", "reviews"]
        cols_2 = ["preco", "preco_original", "rating", "desconto_percentual"]
        cols_3 = ["indice_preco", "score_relevancia", "score_visibilidade", "score_qualidade", "score_atratividade"]

        for col in cols_int:
            if col in header_to_col:
                c = header_to_col[col]
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=c).number_format = fmt_int

        for col in cols_2:
            if col in header_to_col:
                c = header_to_col[col]
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=c).number_format = fmt_2

        for col in cols_3:
            if col in header_to_col:
                c = header_to_col[col]
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=c).number_format = fmt_3

        # Largura simples de colunas
        for idx, h in enumerate(headers, start=1):
            letter = get_column_letter(idx)
            base = 18
            if isinstance(h, str):
                if len(h) <= 8:
                    base = 14
                elif len(h) >= 28:
                    base = 32
            ws.column_dimensions[letter].width = base

    return True




def _deduplicar_resultados(resultados: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Remove duplicados com base em:
    - product_id, quando disponível
    - senão, (produto, seller, preco) como chave aproximada.
    """
    vistos = set()
    filtrados: List[Dict[str, Any]] = []

    for r in resultados:
        product_id = r.get("product_id")
        if product_id:
            chave: Tuple[Any, ...] = ("id", product_id)
        else:
            chave = (
                "fallback",
                r.get("produto"),
                r.get("seller"),
                r.get("preco"),
            )

        if chave in vistos:
            continue

        vistos.add(chave)
        filtrados.append(r)

    return filtrados


def coletar_varias_paginas(
    palavra_chave: str,
    n_paginas: int = N_PAGINAS_DEFAULT,
) -> List[Dict[str, Any]]:
    """
    Consulta o Google Shopping via SearchAPI em múltiplas páginas
    e retorna uma lista consolidada de itens já “parseados”.
    """
    todos_resultados: List[Dict[str, Any]] = []

    for page in range(1, n_paginas + 1):
        print(f"Buscando página {page}/{n_paginas} para '{palavra_chave}'...")

        try:
            resposta_json = buscar_google_shopping(
                q=palavra_chave,
                page=page,
                condition="new",  # ou simplesmente remover essa linha
            )
        except GoogleShoppingError as e:
            print(f"Erro na página {page}: {e}")
            # Estratégia conservadora: para evitar custo, paramos a paginação.
            break

        resultados_pagina = extrair_resultados_basicos(
            resposta_json=resposta_json,
            palavra_chave=palavra_chave,
        )

        # Padronização: garantir 'pagina' nos itens (para bater com Amazon)
        for r in resultados_pagina:
            if isinstance(r, dict):
                r.setdefault("pagina", page)

        if not resultados_pagina:
            print(f"Nenhum resultado retornado na página {page}. Encerrando paginação.")
            break

        todos_resultados.extend(resultados_pagina)

    return todos_resultados


def _adicionar_metricas_relevancia(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adiciona colunas de análise e score de relevância ao DataFrame:

    - preco_outlier (bool)
    - indice_preco
    - desconto_percentual
    - score_preco
    - score_promo
    - score_atratividade
    - score_visibilidade
    - score_qualidade
    - score_relevancia
    """

    df = df.copy()

    # --- 1. Garantir que preço e preço_original são numéricos ---
    df["preco"] = pd.to_numeric(df.get("preco"), errors="coerce")
    df["preco_original"] = pd.to_numeric(df.get("preco_original"), errors="coerce")

    # --- 2. Detecção de outliers de preço (IQR) ---
    valid_prices = df["preco"].dropna()
    df["preco_outlier"] = False

    if len(valid_prices) >= 4:
        q1 = valid_prices.quantile(0.25)
        q3 = valid_prices.quantile(0.75)
        iqr = q3 - q1
        if iqr > 0:
            limite_inferior = q1 - 1.5 * iqr
            limite_superior = q3 + 1.5 * iqr
            mask_outlier = (df["preco"] < limite_inferior) | (df["preco"] > limite_superior)
            df.loc[mask_outlier, "preco_outlier"] = True

    # --- 3. Índice de preço vs mediana ---
    base_mediana = df.loc[~df["preco_outlier"] & df["preco"].notna(), "preco"]
    if base_mediana.empty:
        base_mediana = valid_prices

    preco_mediana = base_mediana.median() if not base_mediana.empty else None

    if preco_mediana and preco_mediana > 0:
        df["indice_preco"] = df["preco"] / preco_mediana
    else:
        df["indice_preco"] = pd.NA

    # --- 4. Desconto percentual ---
    cond_desconto_valido = (
        df["preco_original"].notna()
        & df["preco_original"].gt(0)
        & df["preco"].notna()
        & df["preco_original"].gt(df["preco"])
    )

    df["desconto_percentual"] = 0.0
    df.loc[cond_desconto_valido, "desconto_percentual"] = (
        (df.loc[cond_desconto_valido, "preco_original"] - df.loc[cond_desconto_valido, "preco"])
        / df.loc[cond_desconto_valido, "preco_original"]
    )

    # --- 5. Score de preço ---
    TOLERANCIA_PRECO = 0.5

    df["score_preco"] = pd.NA

    cond_idx_valid = df["indice_preco"].notna()
    cond_idx_le_1 = cond_idx_valid & (df["indice_preco"] <= 1)
    cond_idx_gt_1 = cond_idx_valid & (df["indice_preco"] > 1)

    df.loc[cond_idx_le_1, "score_preco"] = 1.0

    df.loc[cond_idx_gt_1, "score_preco"] = (
        1.0 - (df.loc[cond_idx_gt_1, "indice_preco"] - 1.0) / TOLERANCIA_PRECO
    ).clip(lower=0.0)

    # --- 6. Score de promoção ---
    DESCONTO_REF = 0.30

    df["score_promo"] = 0.0
    cond_desc_valid = df["desconto_percentual"].notna() & (df["desconto_percentual"] > 0)

    df.loc[cond_desc_valid, "score_promo"] = (
        df.loc[cond_desc_valid, "desconto_percentual"] / DESCONTO_REF
    ).clip(upper=1.0)

    # --- 7. Score de atratividade comercial ---
    df["score_atratividade"] = pd.NA

    cond_atrativ_valid = df["score_preco"].notna()
    df.loc[cond_atrativ_valid, "score_atratividade"] = (
        0.7 * df.loc[cond_atrativ_valid, "score_preco"].astype(float)
        + 0.3 * df.loc[cond_atrativ_valid, "score_promo"].astype(float)
    )

    # --- 8. Score de visibilidade ---
    df["posicao"] = pd.to_numeric(df.get("posicao"), errors="coerce")
    df["score_visibilidade"] = pd.NA

    posicoes_validas = df["posicao"].dropna()
    if not posicoes_validas.empty:
        p_max = posicoes_validas.max()
        if p_max > 0:
            df.loc[df["posicao"].notna(), "score_visibilidade"] = (
                (p_max - df.loc[df["posicao"].notna(), "posicao"] + 1.0) / p_max
            )

    # --- 9. Score de qualidade ---
    df["rating"] = pd.to_numeric(df.get("rating"), errors="coerce")
    df["reviews"] = pd.to_numeric(df.get("reviews"), errors="coerce")

    df["score_qualidade"] = pd.NA

    R_REF = 100.0

    cond_qual_valid = df["rating"].notna() & df["reviews"].notna() & (df["rating"] > 0) & (df["reviews"] >= 0)

    if cond_qual_valid.any():
        q_rating = df.loc[cond_qual_valid, "rating"] / 5.0

        def _calc_f_reviews(rev: float) -> float:
            try:
                return min(1.0, math.log(1.0 + rev) / math.log(1.0 + R_REF))
            except (ValueError, ZeroDivisionError):
                return 0.0

        f_reviews = df.loc[cond_qual_valid, "reviews"].apply(_calc_f_reviews)

        df.loc[cond_qual_valid, "score_qualidade"] = (q_rating * f_reviews).clip(upper=1.0)

    # --- 10. Score final de relevância ---
    df["score_relevancia"] = pd.NA

    cond_final = (
        df["score_visibilidade"].notna()
        | df["score_qualidade"].notna()
        | df["score_atratividade"].notna()
    )

    if cond_final.any():
        vis = pd.to_numeric(df.loc[cond_final, "score_visibilidade"], errors="coerce").fillna(0.0)
        qual = pd.to_numeric(df.loc[cond_final, "score_qualidade"], errors="coerce").fillna(0.0)
        atr = pd.to_numeric(df.loc[cond_final, "score_atratividade"], errors="coerce").fillna(0.0)

        df.loc[cond_final, "score_relevancia"] = 0.4 * vis + 0.3 * qual + 0.3 * atr

    return df


# =========================
# Saída padronizada (Diretoria / Análises)
# =========================

EXEC_COLS_ANALISE = [
    "origem",
    "capturado_em",
    "palavra_chave",
    "pagina",
    "posicao",
    "produto",
    "marca",
    "seller",
    "preco",
    "preco_original",
    "desconto_percentual",
    "rating",
    "reviews",
    "patrocinado",
    "id_item",
    "link",
    "indice_preco",
    "score_relevancia",
    "score_visibilidade",
    "score_qualidade",
    "score_atratividade",
]


def _montar_df_analise(df: pd.DataFrame, *, origem: str, palavra_chave: str) -> pd.DataFrame:
    """Cria um DataFrame padronizado (mesmo layout para Amazon e Google Shopping)."""
    if df.empty:
        return df.copy()

    base = df.copy()

    if "capturado_em" not in base.columns:
        base["capturado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Marca (prioridade: canonizada -> marca -> vazio)
    marca = None
    for c in ("marca_canonica", "marca"):
        if c in base.columns:
            marca = base[c]
            break
    if marca is None:
        marca = pd.NA

    # ID do item (Google = product_id)
    id_item = base["product_id"] if "product_id" in base.columns else pd.NA

    def _col_or_na(col: str):
        return base[col] if col in base.columns else pd.NA

    out = pd.DataFrame({
        "origem": origem,
        "capturado_em": base["capturado_em"],
        "palavra_chave": base["palavra_chave"] if "palavra_chave" in base.columns else palavra_chave,
        "pagina": _col_or_na("pagina"),
        "posicao": _col_or_na("posicao"),
        "produto": _col_or_na("produto"),
        "marca": marca,
        "seller": _col_or_na("seller"),
        "preco": _col_or_na("preco"),
        "preco_original": _col_or_na("preco_original"),
        "desconto_percentual": _col_or_na("desconto_percentual"),
        "rating": _col_or_na("rating"),
        "reviews": _col_or_na("reviews"),
        "patrocinado": _col_or_na("patrocinado"),
        "id_item": id_item,
        "link": _col_or_na("link"),
        "indice_preco": _col_or_na("indice_preco"),
        "score_relevancia": _col_or_na("score_relevancia"),
        "score_visibilidade": _col_or_na("score_visibilidade"),
        "score_qualidade": _col_or_na("score_qualidade"),
        "score_atratividade": _col_or_na("score_atratividade"),
    })

    # desconto em % (para leitura executiva). Se vier como fração 0-1, converte para 0-100.
    if "desconto_percentual" in out.columns:
        try:
            s = pd.to_numeric(out["desconto_percentual"], errors="coerce")
            if s.notna().any() and (s.dropna().max() <= 1.0):
                out["desconto_percentual"] = (s * 100.0)
        except Exception:
            pass

    # Garante ordem e presença das colunas
    for c in EXEC_COLS_ANALISE:
        if c not in out.columns:
            out[c] = pd.NA
    out = out[EXEC_COLS_ANALISE]

    # Tipagem numérica
    num_cols = [
        "pagina", "posicao", "preco", "preco_original", "desconto_percentual",
        "rating", "reviews", "indice_preco",
        "score_relevancia", "score_visibilidade", "score_qualidade", "score_atratividade",
    ]
    for c in num_cols:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")

    return out


def _exportar_csv_analise(df_analise: pd.DataFrame, *, origem: str, slug: str, timestamp: str) -> Path:
    """Exporta CSV PT-BR: separador ';' e decimal ',' (bom para Excel no Brasil)."""
    data_exec = datetime.now().strftime("%Y-%m-%d")
    out_dir = ANALISE_DIR / "csv" / data_exec
    out_dir.mkdir(parents=True, exist_ok=True)

    caminho = out_dir / f"Analise_{origem}_{slug}_{timestamp}.csv"
    df_analise.to_csv(caminho, index=False, sep=";", decimal=",", encoding="utf-8-sig")
    return caminho


def main() -> None:
    print("=== Pesquisa de Mercado - Google Shopping (SearchAPI) ===")
    palavra_chave = input("Digite o produto a ser pesquisado (ex.: 'mop spray', 'jogo de panelas'): ").strip()

    if not palavra_chave:
        print("Nenhuma palavra-chave informada. Encerrando.")
        return

    try:
        entrada_paginas = input(f"Quantas páginas deseja buscar? (padrão = {N_PAGINAS_DEFAULT}): ").strip()
        n_paginas = int(entrada_paginas) if entrada_paginas else N_PAGINAS_DEFAULT
        if n_paginas <= 0:
            n_paginas = N_PAGINAS_DEFAULT
    except ValueError:
        n_paginas = N_PAGINAS_DEFAULT

    resultados = coletar_varias_paginas(
        palavra_chave=palavra_chave,
        n_paginas=n_paginas,
    )

    if not resultados:
        print("Nenhum resultado encontrado ou erro na consulta.")
        return

    # Deduplicar
    resultados_unicos = _deduplicar_resultados(resultados)

    # Converter para DataFrame
    df = pd.DataFrame(resultados_unicos)

    # Coluna de timestamp da captura
    df["capturado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Adiciona métricas e scores de relevância
    df = _adicionar_metricas_relevancia(df)

    # Garante que o diretório de saída existe
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # =========================
    # Saída (padrão executivo)
    # =========================
    data_exec = datetime.now().strftime("%Y-%m-%d")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = _slugify(palavra_chave)
    out_dir = OUTPUT_DIR / "csv" / data_exec
    out_dir.mkdir(parents=True, exist_ok=True)
    nome_arquivo = f"GoogleShopping_{slug}_{timestamp}.csv"
    caminho_arquivo = out_dir / nome_arquivo
    df.to_csv(caminho_arquivo, index=False, encoding="utf-8-sig")

    # Saída padronizada para análises (Diretoria) — PT-BR (;) e decimal (,)
    df_analise = _montar_df_analise(df, origem="GoogleShopping", palavra_chave=palavra_chave)
    out_csv_analise = _exportar_csv_analise(df_analise, origem="GoogleShopping", slug=slug, timestamp=timestamp)
    print(f"Arquivo (padronizado p/ análise): {out_csv_analise}")

    # Pequeno resumo de sanity check
    if "score_relevancia" in df.columns:
        print("\nTop 5 produtos por score de relevância:")
        cols_resumo = [
            "produto",
            "seller",
            "marca",
            "preco",
            "score_relevancia",
            "score_visibilidade",
            "score_qualidade",
            "score_atratividade",
        ]
        print(
            df.sort_values("score_relevancia", ascending=False)[cols_resumo]
            .head(5)
            .to_string(index=False)
        )


if __name__ == "__main__":
    main()
