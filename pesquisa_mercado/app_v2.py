import io
import time
from pathlib import Path
import pandas as pd
import streamlit as st

import database as db
from core_v2 import (aplicar_limpeza, calcular_gbb, carregar_regras, coletar_dados_brutos, classificar_faixa_gbb)
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ==========================================
# CONSTANTES DE DESIGN PREMIUM E GLOBAL
# ==========================================

st.set_page_config(
    page_title="Inteligência B2B - Fluxo Modular", 
    layout="wide",
    initial_sidebar_state="expanded"
)

def apply_custom_css():
    st.markdown("""
        <style>
        .block-container { padding-top: 2rem; max-width: 1400px; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif; }
        h1, h2, h3, h4 { color: #111827; font-weight: 600; }
        .stMarkdown p { color: #4b5563; }
        .stButton button { background-color: #111827 !important; color: #ffffff !important; border-radius: 4px !important; padding: 0.6rem 1.5rem !important; font-weight: 500 !important; border: 1px solid #111827 !important; transition: all 0.15s ease; }
        .stButton button:hover { background-color: #374151 !important; border-color: #374151 !important; }
        .stAlert { border-radius: 6px; }
        .step-box { background-color: #f9fafb; border: 1px solid #e5e7eb; border-radius: 8px; padding: 20px; margin-bottom: 20px; }
        </style>
    """, unsafe_allow_html=True)


def fmt_brl(val: float) -> str:
    if pd.isna(val) or val is None:
        return "N/A"
    return f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def format_excel_worksheet(ws, df):
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    numeric_cols = ["preco", "preco_original", "Preço Mínimo", "Preço Médio Mercado", "Preço Máximo", 
                    "Corte P33 (Better)", "Corte P67 (Best)", "Seu Preço", "Seu Preço Médio BI", "Média Global do Mercado", "Média da Sua Faixa (Target)",
                    "Média Econômica", "Média Intermediária", "Média Premium"]
    
    for idx, col in enumerate(df.columns, 1):
        col_series = df[col].astype(str)
        max_len = max(col_series.map(len).max(), len(str(col))) + 2
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len, 40)
        
        if col in numeric_cols:
            for r in range(2, len(df) + 2):
                ws.cell(row=r, column=idx).number_format = '"R$ "#,##0.00'

# ==============================================
# LOGICA DE NEGOCIO E CRUZAMENTO B2B
# ==============================================

def analisar_insight_estrategico(preco_interno, media, corte_p67):
    """Gera insight de ação com base na posição da empresa contra o mercado (agora pela média)."""
    if pd.isna(preco_interno) or pd.isna(media):
        return "Sem dados para cruzar"
        
    diff_percent = ((preco_interno - media) / media) * 100
    
    if preco_interno > corte_p67:
        return "⚠️ ALERTA: Preço Muito Alto (Avaliar redução para não perder market share)"
    elif diff_percent > 15:
        return "💡 ACIMA DA MÉDIA: Produto premium. Monitorar conversão de vendas."
    elif diff_percent < -15:
        return "🚀 OPORTUNIDADE: Preço muito baixo. Avaliar aumento para ganhar margem."
    else:
        return "✅ COMPETITIVO: Preço alinhado com o mercado central."


def parse_keywords(input_text: str, uploaded_file) -> list:
    kws = []
    if input_text: kws.extend([k.strip() for k in input_text.splitlines() if k.strip()])
    if uploaded_file:
        content = uploaded_file.getvalue().decode("utf-8")
        if uploaded_file.name.endswith(".csv"):
            import csv
            reader = csv.reader(io.StringIO(content))
            for row in reader:
                if row and row[0].strip(): kws.append(row[0].strip())
        else:
            kws.extend([k.strip() for k in content.splitlines() if k.strip()])
    return list(dict.fromkeys(kws))


def carregar_tabela_interna(uploaded_file) -> pd.DataFrame:
    try:
        if uploaded_file.name.endswith('.csv'): df_int = pd.read_csv(uploaded_file)
        else: df_int = pd.read_excel(uploaded_file)
            
        cols_originais = list(df_int.columns)
        cols_lower = [str(c).lower().strip() for c in cols_originais]
        df_int.columns = cols_lower
        
        col_cat = [c for c in cols_lower if "categoria" in c or "palavra" in c]
        col_prod = [c for c in cols_lower if "produto" in c or "sku" in c or "item" in c]
        if not col_cat or not col_prod:
            st.error("A planilha deve conter as colunas 'Categoria' e 'Produto'.")
            return pd.DataFrame()
            
        nome_cat, nome_prod = col_cat[0], col_prod[0]
        
        col_pmv = [c for c in cols_lower if "médio" in c or "medio" in c or "pmv" in c]
        nome_pmv = col_pmv[0] if col_pmv else None
        
        id_vars = [nome_cat, nome_prod]
        if nome_pmv: id_vars.append(nome_pmv)
        
        colunas_preco = [c for c in cols_lower if c not in id_vars and 
                         ("pre" in c or "site" in c or "shopee" in c or "mercado" in c or "ml" in c)]
        if not colunas_preco:
            st.error("Nenhuma coluna de preço detectada.")
            return pd.DataFrame()
            
        df_melted = pd.melt(df_int, id_vars=id_vars, value_vars=colunas_preco, var_name="Canal de Venda", value_name="Seu Preço")
        def clean_price(x):
            if pd.isna(x): return None
            if isinstance(x, (int, float)): return float(x)
            s = str(x).upper().replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
            try: return float(s)
            except: return None
            
        df_melted["Seu Preço"] = df_melted["Seu Preço"].apply(clean_price)
        if nome_pmv:
            df_melted[nome_pmv] = df_melted[nome_pmv].apply(clean_price)
            
        df_melted = df_melted.dropna(subset=["Seu Preço"])
        
        rename_dict = {nome_cat: "Categoria", nome_prod: "Produto Interno"}
        if nome_pmv: rename_dict[nome_pmv] = "Preço Médio de Venda"
        
        df_melted = df_melted.rename(columns=rename_dict)
        df_melted["Canal de Venda"] = df_melted["Canal de Venda"].str.title()
        
        cols_out = ["Categoria", "Produto Interno", "Canal de Venda", "Seu Preço"]
        if nome_pmv: cols_out.insert(3, "Preço Médio de Venda")
        
        return df_melted[cols_out]
    except Exception as e:
        st.error(f"Não foi possível ler seu arquivo: {e}")
        return pd.DataFrame()

# ==============================================
# APLICAÇÃO STREAMLIT PRINCIPAL (3 PASSOS)
# ==============================================

def init_session():
    if "market_data_dict" not in st.session_state:
        st.session_state.market_data_dict = {}
    if "df_resumo_gbb" not in st.session_state:
        st.session_state.df_resumo_gbb = None
    if "df_raw_market" not in st.session_state:
        st.session_state.df_raw_market = None
    if "mercado_processado" not in st.session_state:
        st.session_state.mercado_processado = False
    if "df_interno" not in st.session_state:
        st.session_state.df_interno = pd.DataFrame()

def main():
    apply_custom_css()
    init_session()
    
    st.sidebar.markdown("### 🗄️ Integração Power BI (ODS)")
    st.sidebar.caption("Baixe as tabelas do Histórico Completo de Extrações salvos na sua base local, preservando o momento exato em que a pesquisa ocorreu.")
    
    import database as db
    if Path(db.DB_PATH).exists():
        try:
            bytes_db = db.exportar_banco_completo()
            if bytes_db:
                st.sidebar.download_button(
                    label="📥 Exportar Base Master PBI",
                    data=bytes_db,
                    file_name="Master_Historico_Pricing_PBI.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
        except Exception:
            pass
            
    st.title("Inteligência B2B - Pricing")
    st.markdown("Bem-vindo ao fluxo modular em 3 etapas para pesquisa e cruzamento analítico.")
    st.divider()
    
    # PASSO 1
    st.markdown("<div class='step-box'>", unsafe_allow_html=True)
    st.markdown("### Etapa 1: Extração e Definição da Régua (Mercado)")
    
    c1, c2 = st.columns([2, 1])
    with c1:
        input_text = st.text_area("Quais palavras-chave buscar no Google Shopping? (Uma por linha)", height=100)
    with c2:
        uploaded_kws = st.file_uploader("Upload de Txt/Csv de Palavras-Chave", type=["csv", "txt"])
        btn_mercado = st.button("Buscar Mercado e Gerar Régua", use_container_width=True)
    
    if btn_mercado:
        palavras_chave = parse_keywords(input_text, uploaded_kws)
        if not palavras_chave:
            st.error("Insira as palavras-chave para o Passo 1.")
        else:
            barra = st.progress(0)
            status_texto = st.empty()
            regras = carregar_regras()
            
            market_dict = {}
            lista_resumo = []
            todos_brutos = []
            
            for idx, kw in enumerate(palavras_chave):
                kw_nome = str(kw).strip()
                status_texto.markdown(f"**Buscando:** `{kw_nome}`...")
                df_b = coletar_dados_brutos(kw_nome, n_paginas=1)
                if not df_b.empty:
                    df_t, _ = aplicar_limpeza(df_b, regras, kw_nome)
                    if not df_t.empty:
                        gbb = calcular_gbb(df_t)
                        if gbb:
                            df_t["Banda_GBB"] = df_t["preco"].apply(lambda p: classificar_faixa_gbb(p, gbb["corte_p33"], gbb["corte_p67"]))
                            market_dict[kw_nome] = gbb
                            lista_resumo.append({
                                "Categoria": kw_nome,
                                "Volume Base Mercado": gbb["amostra_base"],
                                "Preço Mínimo": gbb["minimo"],
                                "Corte P33 (Better)": gbb["corte_p33"],
                                "Preço Médio Mercado": gbb["media"],
                                "Corte P67 (Best)": gbb["corte_p67"],
                                "Preço Máximo": gbb["maximo"],
                                "Média Econômica": gbb.get("media_economica", gbb["media"]),
                                "Média Intermediária": gbb.get("media_intermediaria", gbb["media"]),
                                "Média Premium": gbb.get("media_premium", gbb["media"]),
                                "Faixa Econômica (Good)": f"De R$ {fmt_brl(gbb['minimo'])} a R$ {fmt_brl(gbb['corte_p33'])}",
                                "Faixa Média (Better)": f"De R$ {fmt_brl(gbb['corte_p33'])} a R$ {fmt_brl(gbb['corte_p67'])}",
                                "Faixa Premium (Best)": f"De R$ {fmt_brl(gbb['corte_p67'])} a R$ {fmt_brl(gbb['maximo'])}",
                            })
                            todos_brutos.append(df_t)
                barra.progress((idx + 1) / len(palavras_chave))
                
            if market_dict:
                st.session_state.market_data_dict = market_dict
                st.session_state.df_resumo_gbb = pd.DataFrame(lista_resumo)
                st.session_state.df_raw_market = pd.concat(todos_brutos, ignore_index=True)
                st.session_state.mercado_processado = True
                
                # Persistir no Histórico ODS do Banco Local
                db.salvar_mercado(st.session_state.df_raw_market, st.session_state.df_resumo_gbb)
                
                status_texto.success("Mercado base estabelecido! Siga para a Etapa 2.")
            else:
                status_texto.error("Não houve retorno da API para gerar o mercado base.")
    
    if st.session_state.mercado_processado:
        st.success(f"Mercado gerado para {len(st.session_state.market_data_dict)} categorias.")
    st.markdown("</div>", unsafe_allow_html=True)
    
    
    # PASSO 2
    if st.session_state.mercado_processado:
        st.markdown("<div class='step-box'>", unsafe_allow_html=True)
        st.markdown("### Etapa 2: Upload da Tabela Mestra (Seu Portfólio)")
        st.markdown("Envie a planilha de itens. Ela deve conter colunas com os nomes `Categoria`, `Produto` e seus canais de preço (ex: `Preço Site`, `Preço ML`).")
        
        uploaded_interno = st.file_uploader("Upload Planilha Interna (Itens)", type=["csv", "xlsx", "xls"])
        if uploaded_interno:
            df_int = carregar_tabela_interna(uploaded_interno)
            if not df_int.empty:
                st.session_state.df_interno = df_int
                st.success(f"Carregados {len(df_int)} preços/itens únicos. Siga para a Etapa 3.")
        st.markdown("</div>", unsafe_allow_html=True)
        
        
    # PASSO 3
    if not st.session_state.df_interno.empty and st.session_state.mercado_processado:
        st.markdown("<div class='step-box'>", unsafe_allow_html=True)
        st.markdown("### Etapa 3: Cruzamento (Relatório Baseado em Oportunidade por Média)")
        
        if st.button("Cruzar Itens e Gerar Excel (Média)", use_container_width=True):
            lista_powerbi = []
            df_interno = st.session_state.df_interno
            market_dict = st.session_state.market_data_dict
            
            # Loop pelo portfolio cruzando com dict em memória
            for _, r in df_interno.iterrows():
                cat = r["Categoria"]
                if cat in market_dict:
                    gbb = market_dict[cat]
                    preco = r["Seu Preço"]
                    preco_ref = r["Preço Médio de Venda"] if "Preço Médio de Venda" in r and not pd.isna(r["Preço Médio de Venda"]) else preco
                    posicionamento = classificar_faixa_gbb(preco_ref, gbb["corte_p33"], gbb["corte_p67"])
                    
                    if "Econômica" in posicionamento: media_alvo = gbb.get("media_economica", gbb.get("media", 0))
                    elif "Premium" in posicionamento: media_alvo = gbb.get("media_premium", gbb.get("media", 0))
                    else: media_alvo = gbb.get("media_intermediaria", gbb.get("media", 0))
                    
                    row_dict = {
                        "Categoria": cat,
                        "Produto Interno": r["Produto Interno"],
                        "Canal de Venda": r["Canal de Venda"],
                        "Seu Preço": preco,
                    }
                    if "Preço Médio de Venda" in r:
                        row_dict["Seu Preço Médio BI"] = r["Preço Médio de Venda"]
                        
                    row_dict["Faixa Praticada (Posicionamento)"] = posicionamento
                    row_dict["Média Global do Mercado"] = gbb["media"]
                    row_dict["Média da Sua Faixa (Target)"] = media_alvo
                    
                    lista_powerbi.append(row_dict)
            
            if not lista_powerbi:
                st.error("Nenhuma categoria da sua planilha internou 'match' com o mercado rastreado no Passo 1.")
            else:
                df_powerbi_final = pd.DataFrame(lista_powerbi)
                
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df_powerbi_final.to_excel(writer, index=False, sheet_name="Fato_Cruzamento_PBI")
                    format_excel_worksheet(writer.sheets["Fato_Cruzamento_PBI"], df_powerbi_final)
                    
                    st.session_state.df_resumo_gbb.to_excel(writer, index=False, sheet_name="Dim_Mercado_GBB")
                    format_excel_worksheet(writer.sheets["Dim_Mercado_GBB"], st.session_state.df_resumo_gbb)
                        
                    st.session_state.df_raw_market.to_excel(writer, index=False, sheet_name="Raw_Market_Data")
                    format_excel_worksheet(writer.sheets["Raw_Market_Data"], st.session_state.df_raw_market)
                val_excel = output.getvalue()
                
                # Persistir no Banco Local a Tabela Fato de Cruzamentos
                db.salvar_cruzamento(df_powerbi_final)
                
                st.success("Cruzamento realizado utilizando as **Médias de Mercado**.")
                st.download_button("📥 Baixar Relatório Final Cruzado (.xlsx)", data=val_excel, file_name=f"Resultado_Mercado_V4_{int(time.time())}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                
        st.markdown("</div>", unsafe_allow_html=True)

if __name__ == "__main__":
    main()
